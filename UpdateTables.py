# Allow the ability to import the excel workbook to work with tables
from openpyxl import load_workbook
import os

# Local location of my file and known titles
FILENAME = r"C:\Users\trevo\OneDrive\Revamped Trev and Taylor Movie Club.xlsx"
CACHED_TITLES = "known_movie_titles"

def load_known_titles():
    if os.path.exists(CACHED_TITLES):
        with open(CACHED_TITLES, "r") as known_titles:  # Opens the file in read mode ("r")
            return set(line.strip() for line in known_titles)  # return a set of the movie titles, strip any extra white space
    return set()  # Returns an empty set if nothing
    
def save_known_titles(titles):
    with open(CACHED_TITLES, "w") as saved_titles:  # Opens file in write mode ("w")
        for title in sorted(titles): # Loops through every title in the parameter set "titles", title is esentially the index, also sorts titles alphabetically in the saved file
            saved_titles.write(f"{title}\n")

def get_sheet_rows(sheet):
    return list(sheet.iter_rows(min_row = 3, values_only = True))  # Iterate through all table rows and grab the values, skip the first two rows as there are headers

    
def main():
    # Get the excel file and store the two pages separately
    excel_doc = load_workbook(FILENAME)
    TnT_sheet = excel_doc[r"Movie Club - TnT"]
    MN_sheet = excel_doc[r"Movie Club - MN"]

    # Get the known titles from previous runs
    known_titles = load_known_titles()

    # Get all info from all rows from each sheet's table
    tnt_rows = get_sheet_rows(TnT_sheet)
    mn_rows  = get_sheet_rows(MN_sheet)

    tnt_titles = set()
    for row in tnt_rows:
        title = row[0]  # first column = title
        if title:       # skip empty titles
            tnt_titles.add(title)

    mn_titles = set()
    for row in mn_rows:
        title = row[0]
        if title:
            mn_titles.add(title)
    
    # 🔍 Find titles that were added to one sheet, and not in the other, or in the cached titles
    tnt_titles_added = tnt_titles - mn_titles - known_titles
    mn_titles_added = mn_titles - tnt_titles - known_titles

    # Add titles found in TnT and not MN, to MN
    for row in tnt_rows:
        if row[0] in tnt_titles_added: # row[0] is checking the indexed row, column "0" which is column 1, the movie title
            new_row = list(row[:4]) + ["", "", "", "", ""]
            MN_sheet.append(new_row)
            
    # Add titles found in MN and not Tnt, to Tnt
    for row in mn_rows:
        if row[0] in mn_titles_added:
            new_row = list(row[:4]) + ["", "", "", "", ""]
            TnT_sheet.append(new_row)

    # If any new movies were added, save them to the .txt file
    if tnt_titles_added or mn_titles_added:
        excel_doc.save(FILENAME)
        all_titles = tnt_titles.union(mn_titles)
        save_known_titles(all_titles)

        print("✅ Synced new titles between sheets:")
        if tnt_titles_added:
            print(" - From Sheet1 to Sheet2:", tnt_titles_added)
        if mn_titles_added:
            print(" - From Sheet2 to Sheet1:", mn_titles_added)
    else:
        print("✅ No new titles to sync.")

if __name__ == "__main__":
    main()